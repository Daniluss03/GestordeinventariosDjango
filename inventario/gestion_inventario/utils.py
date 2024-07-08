# En gestion_inventario/utils.py

from django.conf import settings
import os
import matplotlib.pyplot as plt
from .models import Articulo

def generar_grafica_productos_mas_vendidos_por_categoria():
    # Obtener categorías únicas
    categorias = Articulo.CATEGORIAS

    # Generar gráfica para cada categoría
    for categoria_display, _ in categorias:
        articulos = Articulo.objects.filter(categoria=categoria_display).order_by('-ventas')[:5]  # Tomar los 5 más vendidos por categoría
        nombres = [articulo.nombre for articulo in articulos]
        ventas = [articulo.ventas for articulo in articulos]

        # Crear la gráfica
        plt.figure(figsize=(10, 6))
        plt.bar(nombres, ventas, color='blue')
        plt.xlabel('Artículos')
        plt.ylabel('Ventas')
        plt.title(f'Productos más vendidos - {categoria_display}')

        # Guardar la gráfica en formato PNG en el directorio de archivos estáticos
        static_dir = os.path.join(settings.BASE_DIR, 'static')
        graficas_dir = os.path.join(static_dir, 'graficas')
        if not os.path.exists(graficas_dir):
            os.makedirs(graficas_dir)

        filename = f'productos_mas_vendidos_{categoria_display.lower().replace(" ", "_")}.png'
        path = os.path.join(graficas_dir, filename)
        plt.savefig(path)

        # Limpiar la figura para liberar recursos
        plt.clf()

def generar_grafica_productos_menos_vendidos_por_categoria():
    # Obtener categorías únicas
    categorias = Articulo.CATEGORIAS

    # Generar gráfica para cada categoría
    for categoria_display, _ in categorias:
        articulos = Articulo.objects.filter(categoria=categoria_display).order_by('ventas')[:5]  # Tomar los 5 menos vendidos por categoría
        nombres = [articulo.nombre for articulo in articulos]
        ventas = [articulo.ventas for articulo in articulos]

        # Crear la gráfica
        plt.figure(figsize=(10, 6))
        plt.pie(ventas, labels=nombres, autopct='%1.1f%%', startangle=140)
        plt.title(f'Productos menos vendidos - {categoria_display}')

        # Guardar la gráfica en formato PNG en el directorio de archivos estáticos
        static_dir = os.path.join(settings.BASE_DIR, 'static')
        graficas_dir = os.path.join(static_dir, 'graficas')
        if not os.path.exists(graficas_dir):
            os.makedirs(graficas_dir)

        filename = f'productos_menos_vendidos_{categoria_display.lower().replace(" ", "_")}.png'
        path = os.path.join(graficas_dir, filename)
        plt.savefig(path)

        # Limpiar la figura para liberar recursos
        plt.clf()

def generar_grafica_productos_mas_entradas_por_categoria():
    # Obtener categorías únicas
    categorias = Articulo.CATEGORIAS

    # Generar gráfica para cada categoría
    for categoria_display, _ in categorias:
        articulos = Articulo.objects.filter(categoria=categoria_display).order_by('-entradas')[:5]  # Tomar los 5 con más entradas por categoría
        nombres = [articulo.nombre for articulo in articulos]
        entradas = [articulo.entradas for articulo in articulos]

        # Crear la gráfica
        plt.figure(figsize=(10, 6))
        plt.plot(nombres, entradas, marker='o', linestyle='-', color='green')
        plt.xlabel('Artículos')
        plt.ylabel('Entradas')
        plt.title(f'Productos con más entradas - {categoria_display}')

        # Guardar la gráfica en formato PNG en el directorio de archivos estáticos
        static_dir = os.path.join(settings.BASE_DIR, 'static')
        graficas_dir = os.path.join(static_dir, 'graficas')
        if not os.path.exists(graficas_dir):
            os.makedirs(graficas_dir)

        filename = f'productos_mas_entradas_{categoria_display.lower().replace(" ", "_")}.png'
        path = os.path.join(graficas_dir, filename)
        plt.savefig(path)

        # Limpiar la figura para liberar recursos
        plt.clf()

def generar_grafica_productos_menos_entradas_por_categoria():
    # Obtener categorías únicas
    categorias = Articulo.CATEGORIAS

    # Generar gráfica para cada categoría
    for categoria_display, _ in categorias:
        articulos = Articulo.objects.filter(categoria=categoria_display).order_by('entradas')[:5]  # Tomar los 5 con menos entradas por categoría
        nombres = [articulo.nombre for articulo in articulos]
        entradas = [articulo.entradas for articulo in articulos]

        # Crear la gráfica
        plt.figure(figsize=(10, 6))
        plt.plot(nombres, entradas, marker='o', linestyle='-', color='orange')
        plt.xlabel('Artículos')
        plt.ylabel('Entradas')
        plt.title(f'Productos con menos entradas - {categoria_display}')

        # Guardar la gráfica en formato PNG en el directorio de archivos estáticos
        static_dir = os.path.join(settings.BASE_DIR, 'static')
        graficas_dir = os.path.join(static_dir, 'graficas')
        if not os.path.exists(graficas_dir):
            os.makedirs(graficas_dir)

        filename = f'productos_menos_entradas_{categoria_display.lower().replace(" ", "_")}.png'
        path = os.path.join(graficas_dir, filename)
        plt.savefig(path)

        # Limpiar la figura para liberar recursos
        plt.clf()

import openpyxl
from .models import Articulo

def generar_excel_articulos():
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artículos"

    # Escribir encabezados
    ws['A1'] = "Nombre"
    ws['B1'] = "Cantidad"
    ws['C1'] = "Precio"
    ws['D1'] = "Ventas"
    ws['E1'] = "Entradas"
    ws['F1'] = "Fecha de registro"

    # Obtener todos los artículos desde la base de datos
    articulos = Articulo.objects.all()

    # Escribir datos de los artículos en el archivo Excel
    for index, articulo in enumerate(articulos, start=2):  # Empezar desde la fila 2
        ws[f'A{index}'] = articulo.nombre
        ws[f'B{index}'] = articulo.cantidad
        ws[f'C{index}'] = articulo.precio
        ws[f'D{index}'] = articulo.ventas
        ws[f'E{index}'] = articulo.entradas
        ws[f'F{index}'] = articulo.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')

    # Guardar el archivo Excel en el directorio de archivos estáticos
    excel_path = 'static/articulos.xlsx'
    wb.save(excel_path)

    return excel_path